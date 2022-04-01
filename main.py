import pandas as pd
import openpyxl
import demoji

df = pd.read_excel('Data_Sentiment.xlsx')

data_list = []
sentiment_list = []

wb = openpyxl.load_workbook("Data_Sentiment.xlsx")
ws = wb['Sheet1']
colomn = ws["A"]

for i in range(2,len(colomn)):
    C = 'A' + str(i)
    D = 'B' + str(i)

    x = ws[C].value
    y = ws[D].value

    punc = '''!()-[]{};:'"\,<>„".…|/✔।!,??@#ঃ$%^&*_~।'''
    line = str(x)
    for ele in line:
        if ele in punc:
            line = line.replace(ele, "")
    line = line.replace("\n", ' ').replace("\t", ' ').replace("   ", ' ').replace("  ", ' ')
    line = "".join(i for i in line if i in ["।"] or 2432 <= ord(i) <= 2559 or ord(i) == 32)
    line = " ".join(line.split())
    line = demoji.replace(line, "")

    data_list.append(line)
    sentiment_list.append(y)


print(len(data_list))
print(len(sentiment_list))

wb_1 = openpyxl.load_workbook('clean_dataset.xlsx')
ws_1 = wb_1['Sheet1']
for j in range(2,len(sentiment_list)):
    E = 'A' + str(j)
    F = 'B'+ str(j)
    ws_1[E].value = data_list[j-2]
    ws_1[F].value = sentiment_list[j-2]

wb_1.save("clean_dataset.xlsx")